"""
Utility functions for Asset Management Platform
"""
import os
import io
from datetime import datetime, date
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from sqlalchemy.orm import Session
from models import Asset, AssetOperation, Company
import logging

logger = logging.getLogger(__name__)

def generate_inventory_number() -> str:
    """
    Generate unique inventory number in format: INV-YYYYMMDD-XXXX
    """
    today = datetime.now()
    date_part = today.strftime("%Y%m%d")
    
    # In production, you might want to use database sequence or Redis counter
    # For now, using timestamp-based approach
    time_part = today.strftime("%H%M")
    
    return f"INV-{date_part}-{time_part}"

def generate_unique_inventory_number(db: Session, company_id: int) -> str:
    """
    Generate truly unique inventory number by checking database
    """
    attempts = 0
    max_attempts = 100
    
    while attempts < max_attempts:
        inv_number = generate_inventory_number()
        
        # Check if number exists in this company
        existing = db.query(Asset).join(Asset.warehouse).join(Asset.warehouse.branch).filter(
            Asset.inventory_number == inv_number,
            Asset.warehouse.branch.company_id == company_id
        ).first()
        
        if not existing:
            return inv_number
        
        attempts += 1
        
    # Fallback with random suffix
    import random
    suffix = str(random.randint(1000, 9999))
    today = datetime.now().strftime("%Y%m%d")
    return f"INV-{today}-{suffix}"

def format_currency(amount: float, currency: str = "₽") -> str:
    """Format currency with proper separators"""
    return f"{currency}{amount:,.2f}"

def format_date(dt: datetime, format_str: str = "%d.%m.%Y") -> str:
    """Format date for display"""
    if dt is None:
        return ""
    return dt.strftime(format_str)

def format_datetime(dt: datetime, format_str: str = "%d.%m.%Y %H:%M") -> str:
    """Format datetime for display"""
    if dt is None:
        return ""
    return dt.strftime(format_str)

class ExcelExporter:
    """Excel export utility class"""
    
    def __init__(self):
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        
    def create_assets_report(self, assets: List[Asset], company_name: str) -> io.BytesIO:
        """Create Excel report for assets"""
        
        # Set up worksheet
        self.worksheet.title = "Активы"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Company header
        self.worksheet.merge_cells('A1:L1')
        company_cell = self.worksheet['A1']
        company_cell.value = f"Отчет по активам - {company_name}"
        company_cell.font = Font(bold=True, size=14)
        company_cell.alignment = Alignment(horizontal="center")
        
        # Date header
        self.worksheet.merge_cells('A2:L2')
        date_cell = self.worksheet['A2']
        date_cell.value = f"Сгенерирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        date_cell.alignment = Alignment(horizontal="center")
        
        # Column headers
        headers = [
            "Инвентарный номер",
            "Название",
            "Категория",
            "Статус",
            "Количество",
            "Стоимость",
            "Общая стоимость",
            "Склад",
            "Серийный номер",
            "Поставщик",
            "Дата покупки",
            "Примечания"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        total_value = 0
        for row, asset in enumerate(assets, 5):
            asset_total = asset.cost * asset.quantity
            total_value += asset_total
            
            data = [
                asset.inventory_number,
                asset.name,
                asset.category.value,
                self._get_status_text(asset.status.value),
                asset.quantity,
                asset.cost,
                asset_total,
                asset.warehouse.name if asset.warehouse else "",
                asset.serial_number or "",
                asset.supplier or "",
                format_date(asset.purchase_date) if asset.purchase_date else "",
                asset.notes or ""
            ]
            
            for col, value in enumerate(data, 1):
                cell = self.worksheet.cell(row=row, column=col)
                cell.value = value
                
                # Format currency columns
                if col in [6, 7]:  # Cost and total cost columns
                    cell.number_format = '#,##0.00'
        
        # Summary row
        summary_row = len(assets) + 6
        self.worksheet.merge_cells(f'A{summary_row}:F{summary_row}')
        summary_cell = self.worksheet[f'A{summary_row}']
        summary_cell.value = f"Итого активов: {len(assets)}"
        summary_cell.font = Font(bold=True)
        
        total_cell = self.worksheet[f'G{summary_row}']
        total_cell.value = total_value
        total_cell.font = Font(bold=True)
        total_cell.number_format = '#,##0.00'
        
        # Auto-adjust column widths
        self._auto_adjust_columns()
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        self.workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
    
    def create_operations_report(self, operations: List[AssetOperation], company_name: str) -> io.BytesIO:
        """Create Excel report for operations"""
        
        self.worksheet.title = "Операции"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Company header
        self.worksheet.merge_cells('A1:J1')
        company_cell = self.worksheet['A1']
        company_cell.value = f"Отчет по операциям - {company_name}"
        company_cell.font = Font(bold=True, size=14)
        company_cell.alignment = Alignment(horizontal="center")
        
        # Date header
        self.worksheet.merge_cells('A2:J2')
        date_cell = self.worksheet['A2']
        date_cell.value = f"Сгенерирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        date_cell.alignment = Alignment(horizontal="center")
        
        # Column headers
        headers = [
            "Дата операции",
            "Тип операции",
            "Актив",
            "Количество",
            "Откуда",
            "Куда",
            "Пользователь",
            "Причина",
            "Документ",
            "Примечания"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        for row, operation in enumerate(operations, 5):
            data = [
                format_datetime(operation.operation_date),
                self._get_operation_type_text(operation.type.value),
                operation.asset.name if operation.asset else "",
                operation.quantity,
                operation.from_warehouse.name if operation.from_warehouse else "Внешний",
                operation.to_warehouse.name if operation.to_warehouse else "Внешний",
                operation.user.username if operation.user else "",
                operation.reason or "",
                operation.document_number or "",
                operation.notes or ""
            ]
            
            for col, value in enumerate(data, 1):
                self.worksheet.cell(row=row, column=col, value=value)
        
        # Summary row
        summary_row = len(operations) + 6
        self.worksheet.merge_cells(f'A{summary_row}:D{summary_row}')
        summary_cell = self.worksheet[f'A{summary_row}']
        summary_cell.value = f"Всего операций: {len(operations)}"
        summary_cell.font = Font(bold=True)
        
        # Auto-adjust column widths
        self._auto_adjust_columns()
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        self.workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
    
    def _get_status_text(self, status: str) -> str:
        """Convert status to Russian text"""
        status_map = {
            "Active": "Активен",
            "Inactive": "Неактивен",
            "Repair": "Ремонт",
            "Disposed": "Списан"
        }
        return status_map.get(status, status)
    
    def _get_operation_type_text(self, op_type: str) -> str:
        """Convert operation type to Russian text"""
        type_map = {
            "Receipt": "Поступление",
            "Transfer": "Перемещение",
            "Disposal": "Списание",
            "Adjustment": "Корректировка"
        }
        return type_map.get(op_type, op_type)
    
    def _auto_adjust_columns(self):
        """Auto-adjust column widths"""
        for column in self.worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            self.worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

def create_backup_filename(prefix: str = "backup") -> str:
    """Create timestamped backup filename"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{timestamp}.sql"

def validate_file_upload(file_content: bytes, allowed_extensions: List[str]) -> bool:
    """Validate uploaded file"""
    # Basic file validation
    if len(file_content) > 10 * 1024 * 1024:  # 10MB limit
        return False
    
    # Add more validation as needed
    return True

def sanitize_filename(filename: str) -> str:
    """Sanitize filename for safe storage"""
    import re
    # Remove or replace unsafe characters
    filename = re.sub(r'[^\w\s-.]', '', filename)
    filename = re.sub(r'[-\s]+', '-', filename)
    return filename.strip('-.')

def calculate_depreciation(cost: float, purchase_date: datetime, useful_life_years: int = 5) -> Dict[str, Any]:
    """Calculate asset depreciation"""
    if not purchase_date:
        return {"annual_depreciation": 0, "accumulated_depreciation": 0, "book_value": cost}
    
    years_since_purchase = (datetime.now() - purchase_date).days / 365.25
    annual_depreciation = cost / useful_life_years
    accumulated_depreciation = min(annual_depreciation * years_since_purchase, cost)
    book_value = max(cost - accumulated_depreciation, 0)
    
    return {
        "annual_depreciation": round(annual_depreciation, 2),
        "accumulated_depreciation": round(accumulated_depreciation, 2),
        "book_value": round(book_value, 2),
        "years_since_purchase": round(years_since_purchase, 2)
    }

# Logging utility
def log_audit_action(
    user_id: int,
    company_id: int,
    action: str,
    resource_type: str,
    resource_id: int = None,
    old_values: Dict = None,
    new_values: Dict = None,
    db: Session = None
):
    """Log audit action (simplified - in production use proper audit system)"""
    logger.info(
        f"AUDIT: User {user_id} (Company {company_id}) performed {action} on {resource_type} {resource_id}"
    )